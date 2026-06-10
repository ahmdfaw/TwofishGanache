from flask import Flask, render_template, request, jsonify, send_file
import json
import hashlib
import os
import time  
from datetime import datetime
from web3 import Web3
from twofish_manual import Twofish
import openpyxl

app = Flask(__name__)

# ==========================================
# KONFIGURASI BLOCKCHAIN
# ==========================================
ganache_url = "http://127.0.0.1:7545"
web3 = Web3(Web3.HTTPProvider(ganache_url))

# CONTRACT ADDRESS DARI TRUFFLE MIGRATE
CONTRACT_ADDRESS = "0xE2bEFB28b75dB37C3ad6053C4dB65d4e31c75A2D" 

with open('blockchain/build/contracts/DocumentRegistry.json', 'r') as file:
    contract_json = json.load(file)
    contract_abi = contract_json['abi']

contract = web3.eth.contract(address=CONTRACT_ADDRESS, abi=contract_abi)
web3.eth.default_account = web3.eth.accounts[0]

HISTORY_FILE = 'history.json'

# ==========================================
# FUNGSI MANAJEMEN RIWAYAT (JSON LOKAL)
# ==========================================
def load_history():
    if os.path.exists(HISTORY_FILE):
        try:
            with open(HISTORY_FILE, 'r') as f:
                return json.load(f)
        except:
            return []
    return []

def save_to_history(filename, ciphertext_hex, doc_hash, tx_hash, enc_filename, 
                    waktu_enkripsi, ukuran_plaintext, ukuran_ciphertext, 
                    waktu_penyimpanan, bit_berubah, total_bit):
    history = load_history()
    if not isinstance(history, list):
        history = []
    
    waktu_sekarang = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    new_entry = {
        "timestamp": waktu_sekarang,
        "nama_file": filename,
        "ciphertext": ciphertext_hex,
        "sha_256": doc_hash,
        "tx_hash": tx_hash,
        "enc_filename": enc_filename,
        "waktu_enkripsi": waktu_enkripsi,
        "ukuran_plaintext": ukuran_plaintext,
        "ukuran_ciphertext": ukuran_ciphertext,
        "waktu_penyimpanan": waktu_penyimpanan,
        "bit_berubah": bit_berubah,
        "total_bit": total_bit
    }
    history.append(new_entry)
    with open(HISTORY_FILE, 'w') as f:
        json.dump(history, f, indent=4)

# ==========================================
# FUNGSI UTAMA TWOFISH
# ==========================================
def pad_data(data):
    padding_len = 16 - (len(data) % 16)
    return data + bytes([padding_len] * padding_len)

def unpad_data(data):
    padding_len = data[-1]
    return data[:-padding_len]

def derive_key_sha256(password: str) -> bytes:
    if isinstance(password, str):
        password = password.encode('utf-8')
    return hashlib.sha256(password).digest()

# ==========================================
# RUTE API (FLASK ENDPOINTS)
# ==========================================
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/history', methods=['GET'])
def get_history():
    return jsonify(load_history())

@app.route('/api/encrypt', methods=['POST'])
def api_encrypt():
    try:
        if 'file' not in request.files or 'key' not in request.form:
            return jsonify({"status": "error", "message": "File atau kunci tidak ditemukan"}), 400
            
        file = request.files['file']
        key_string = request.form['key']
        file_bytes = file.read()

        ukuran_plaintext = len(file_bytes)

        key_bytes = derive_key_sha256(key_string)
        T = Twofish(key_bytes)
        padded_data = pad_data(file_bytes)
        
        iv = os.urandom(16)
        ciphertext = iv
        prev_block = iv
        
        start_time = time.time()
        for i in range(0, len(padded_data), 16):
            pt_block = padded_data[i:i+16]
            xored_block = bytes(a ^ b for a, b in zip(pt_block, prev_block))
            ct_block = T.encrypt(xored_block)
            ciphertext += ct_block
            prev_block = ct_block
        duration = time.time() - start_time
        waktu_enkripsi = f"{duration:.4f}"

        ukuran_ciphertext = len(ciphertext)

        actual_ct = ciphertext[16:] 
        total_bit = len(padded_data) * 8
        bit_berubah = sum(bin(b1 ^ b2).count('1') for b1, b2 in zip(padded_data, actual_ct))

        hash_ciphertext = hashlib.sha256(ciphertext).hexdigest()

        enc_filename = f"terenkripsi_{file.filename}.enc"
        enc_filepath = os.path.join('encrypted', enc_filename)
        with open(enc_filepath, 'wb') as f:
            f.write(ciphertext)

        ciphertext_full = ciphertext.hex().upper()

        return jsonify({
            "status": "success",
            "message": "Dokumen berhasil dienkripsi secara lokal.",
            "hash": hash_ciphertext,
            "filename": enc_filename,
            "original_filename": file.filename,
            "ciphertext_full": ciphertext_full,
            "waktu_enkripsi": waktu_enkripsi,
            "ukuran_plaintext": ukuran_plaintext,
            "ukuran_ciphertext": ukuran_ciphertext,
            "bit_berubah": bit_berubah,
            "total_bit": total_bit
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/store_blockchain', methods=['POST'])
def api_store_blockchain():
    try:
        data = request.json
        doc_hash = data.get('hash')
        filename = data.get('filename')
        original_filename = data.get('original_filename')
        ciphertext_full = data.get('ciphertext_full')
        
        waktu_enkripsi = data.get('waktu_enkripsi', '0.0000')
        ukuran_plaintext = data.get('ukuran_plaintext', 0)
        ukuran_ciphertext = data.get('ukuran_ciphertext', 0)
        bit_berubah = data.get('bit_berubah', 0)
        total_bit = data.get('total_bit', 0)

        is_exist = contract.functions.verifyDocument(doc_hash).call()
        if is_exist:
             return jsonify({"status": "error", "message": "Dokumen ini sudah ada di Blockchain!"}), 400

        start_storage_time = time.time()
        tx_hash = contract.functions.storeDocument(doc_hash).transact()
        tx_receipt = web3.eth.wait_for_transaction_receipt(tx_hash)
        storage_duration = time.time() - start_storage_time
        waktu_penyimpanan = f"{storage_duration:.4f}"

        save_to_history(original_filename, ciphertext_full, doc_hash, tx_receipt.transactionHash.hex(), 
                        filename, waktu_enkripsi, ukuran_plaintext, ukuran_ciphertext, 
                        waktu_penyimpanan, bit_berubah, total_bit)

        return jsonify({
            "status": "success",
            "message": "Hash berhasil dikunci di Blockchain",
            "tx_hash": tx_receipt.transactionHash.hex(),
            "waktu_penyimpanan": waktu_penyimpanan
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/decrypt', methods=['POST'])
def api_decrypt():
    try:
        if 'file' not in request.files or 'key' not in request.form:
            return jsonify({"status": "error", "message": "File atau kunci tidak ditemukan"}), 400
            
        file = request.files['file']
        key_string = request.form['key']
        full_ciphertext = file.read()

        hash_ciphertext = hashlib.sha256(full_ciphertext).hexdigest()

        is_exist = contract.functions.verifyDocument(hash_ciphertext).call()
        if not is_exist:
             return jsonify({"status": "error", "message": "DOKUMEN PALSU! Tidak terdaftar di Blockchain."}), 400

        key_bytes = derive_key_sha256(key_string)
        T = Twofish(key_bytes)
        
        iv = full_ciphertext[:16]
        actual_ciphertext = full_ciphertext[16:]
        prev_block = iv
        
        plaintext_padded = b''
        
        start_time = time.time()
        for i in range(0, len(actual_ciphertext), 16):
            ct_block = actual_ciphertext[i:i+16]
            decrypted_block = T.decrypt(ct_block)
            pt_block = bytes(a ^ b for a, b in zip(decrypted_block, prev_block))
            plaintext_padded += pt_block
            prev_block = ct_block
            
        plaintext = unpad_data(plaintext_padded)
        duration = time.time() - start_time
        waktu_dekripsi = f"{duration:.4f}"

        if not plaintext.startswith(b'%PDF-'):
            return jsonify({"status": "error", "message": "Kunci Salah atau File Rusak!"}), 400

        dec_filename = f"asli_{file.filename.replace('.enc', '.pdf')}"
        dec_filepath = os.path.join('decrypted', dec_filename)
        with open(dec_filepath, 'wb') as f:
            f.write(plaintext)

        return jsonify({
            "status": "success",
            "message": "Dokumen valid dan berhasil dibuka!",
            "filename": dec_filename,
            "waktu_dekripsi": waktu_dekripsi 
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

# ==========================================
# RUTE EKSPOR EXCEL TANPA STYLING (VERSI STABIL)
# ==========================================
@app.route('/download_excel/<tx_hash>')
def download_excel(tx_hash):
    history = load_history()
    entry = next((item for item in history if item['tx_hash'] == tx_hash), None)
    
    if not entry:
        return "Data tidak ditemukan", 404

    # Pastikan folder reports ada
    if not os.path.exists('reports'):
        os.makedirs('reports')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hasil Analisis"
    
    bit_plain = entry.get('ukuran_plaintext', 0) * 8
    bit_cipher = entry.get('ukuran_ciphertext', 0) * 8
    bit_compared = entry.get('total_bit', 0)
    changed_bit = entry.get('bit_berubah', 0)
    
    # Hitung persentase langsung (dihindari penggunaan formula Excel)
    pct = (changed_bit / bit_compared) * 100 if bit_compared > 0 else 0
        
    data = [
        ["Metrik Pengujian", "Nilai"],
        ["total bit plain", bit_plain],
        ["total bit cipher", bit_cipher],
        ["total bit compared", bit_compared],
        ["changed bit", changed_bit],
        ["avalanche percent", f"{pct:.2f}%"]
    ]
    
    for row in data:
        ws.append(row)
        
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 16
    
    filename = f"Avalanche_{entry.get('nama_file').replace('.pdf', '')}.xlsx"
    filepath = os.path.join('reports', filename)
    wb.save(filepath)
    wb.close() # Penting: Menutup workbook untuk melepas lock file
    
    return send_file(filepath, as_attachment=True)

# ==========================================
# RUTE EKSPOR SELURUH RIWAYAT (POLOS)
# ==========================================
@app.route('/download_all_history')
def download_all_history():
    history = load_history()
    
    if not history:
        return "Data riwayat kosong", 404

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Riwayat Transaksi"
    
    # Header kolom sesuai permintaan Anda
    headers = [
        "Timestamp", "Nama File", "Size Plaintext (Byte)", "Size Ciphertext (Byte)", 
        "Waktu Enkripsi (s)", "Waktu Penyimpanan (s)", "Avalanche Percent", 
        "Tx Hash", "Changed Bit", "Total Bit"
    ]
    ws.append(headers)
    
    # Mengisi data ke baris-baris berikutnya
    for entry in history:
        bit_berubah = entry.get('bit_berubah', 0)
        total_bit = entry.get('total_bit', 1) # Mencegah pembagian dengan nol
        pct = (bit_berubah / total_bit) * 100 if total_bit > 0 else 0
        
        row = [
            entry.get('timestamp', '-'),
            entry.get('nama_file', '-'),
            entry.get('ukuran_plaintext', 0),
            entry.get('ukuran_ciphertext', 0),
            entry.get('waktu_enkripsi', '0.0000'),
            entry.get('waktu_penyimpanan', '0.0000'),
            f"{pct:.2f}%",
            entry.get('tx_hash', '-'),
            bit_berubah,
            total_bit
        ]
        ws.append(row)
    
    # Atur lebar kolom agar rapi
    for col in range(1, 11):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    
    filename = f"Riwayat_Lengkap_Enkripsi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join('reports', filename)
    wb.save(filepath)
    
    return send_file(filepath, as_attachment=True)

@app.route('/download/<folder>/<filename>')
def download_file(folder, filename):
    if folder not in ['encrypted', 'decrypted']:
        return "Folder tidak valid", 400
    return send_file(os.path.join(folder, filename), as_attachment=True)

if __name__ == '__main__':
    os.makedirs('encrypted', exist_ok=True)
    os.makedirs('decrypted', exist_ok=True)
    os.makedirs('reports', exist_ok=True)
    app.run(debug=True, port=5000)